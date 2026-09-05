#!/usr/bin/env python3
"""Read supplied matching workbook and public Weds catalog; audit every model.

No inferred manufacturer defaults. Raw spreadsheets are never changed.
Cached public pages live outside the repository. Run with bundled Python.
"""
import argparse
import concurrent.futures
import hashlib
import html
import json
import re
import time
import unicodedata
import urllib.request
from pathlib import Path
from urllib.parse import urljoin, quote
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
MAKERS = {'トヨタ':'toyota','レクサス':'lexus','日産':'nissan','ホンダ':'honda','マツダ':'mazda','SUBARU':'subaru','スズキ':'suzuki','ダイハツ':'daihatsu','三菱':'mitsubishi'}
CACHE = Path('/tmp/tire-fitment-public-cache')

def norm(value):
    s = unicodedata.normalize('NFKC', str(value or '')).lower()
    s = ''.join(chr(ord(c)-0x60) if 'ァ' <= c <= 'ヶ' else c for c in s)
    return re.sub(r'[\s\-‐‑‒–—―・_/ー：:（）()]', '', s)

def plain(s):
    return re.sub(r'\s+', ' ', html.unescape(re.sub('<[^>]+>', ' ', s))).strip()

def fetch(url):
    CACHE.mkdir(exist_ok=True)
    path = CACHE / (hashlib.sha256(url.encode()).hexdigest()+'.html')
    if path.exists(): return path.read_text()
    for attempt in range(2):
        try:
            with urllib.request.urlopen(quote(url, safe=':/?=&%'), timeout=25) as r:
                data = r.read().decode('utf-8')
            path.write_text(data)
            return data
        except Exception:
            if attempt: raise
            time.sleep(1)

def index(maker, slug):
    url = f'https://search.weds.co.jp/maker/{slug}/'
    text = fetch(url)
    records = []
    for name, body in re.findall(r'<dt id="[^"]+">(.*?)</dt>(.*?)</dl>', text, re.S):
        for link, label in re.findall(r'<a[^>]*href="([^"]+)"[^>]*>(.*?)</a>', body, re.S):
            records.append(dict(maker=maker, model=plain(name), generation=plain(label), url=urljoin(url,link)))
    return records

def detail(record):
    try:
        landing = fetch(record['url'])
        links = re.findall(r'href="([^"]+)"', landing)
        link = next((x for x in links if x == 'leonis-all'), None)
        if not link: link = next((x for x in links if re.fullmatch(r'[a-z0-9-]+-all',x)), None)
        if not link: return {**record, 'error':'no detail link'}
        url = urljoin(record['url'],link+'/')
        try:
            text = plain(fetch(url))
        except Exception:
            text=''
            for fallback in ['17inch','14inch','12inch','18inch','20inch','16inch']:
                if fallback not in links: continue
                try:
                    url=urljoin(record['url'],fallback)
                    text=plain(fetch(url)); break
                except Exception: pass
            if not text: raise ValueError('no available size detail')
        p = re.search(r'H/P\.C\.D\s+(\d)/(\d+(?:\.\d+)?)', text)
        h = re.search(r'車両ハブ径\s+φ(\d+(?:\.\d+)?)',text)
        f = re.search(r'ボルトサイズ\s+M(\d+)x(\d+(?:\.\d+)?)\s+(ナット|ボルト)',text)
        return {**record, 'source_url':url, 'holes':int(p[1]) if p else None,'pcd':float(p[2]) if p else None,
                'hub_bore':float(h[1]) if h else None,'thread_diameter':int(f[1]) if f else None,
                'thread_pitch':float(f[2]) if f else None,'method':f[3] if f else None}
    except Exception as e: return {**record,'error':str(e)}

def workbook_rows(path):
    wb=load_workbook(path,read_only=True,data_only=True)
    out=[]
    for maker,slug in MAKERS.items():
        sheet=slug.upper()+'－'
        current_name = current_period = current_code = ''
        block=0
        for rownum,row in enumerate(wb[sheet].iter_rows(min_row=5,values_only=True),5):
            name=str(row[0] or '').strip()
            if not name: continue
            if name!=current_name or row[1]:
                block+=1; current_period=''; current_code=''
            current_name=name
            if row[1]: current_period=str(row[1])
            if row[2]: current_code=str(row[2])
            out.append(dict(maker=maker,model=name,period=current_period,code=current_code,drive=str(row[3] or ''),
                tire=str(row[4] or ''),wheel=str(row[5] or ''),pcd=str(row[6] or ''),hub=row[7],fastener=str(row[8] or ''),
                sheet=sheet,row=rownum,block=block))
    return out

def main():
    ap=argparse.ArgumentParser(); ap.add_argument('workbook'); ap.add_argument('--out',default='/tmp/tire-all-fitment-research.json'); args=ap.parse_args()
    master=json.loads((ROOT/'app/data/jp_vehicle_search_master_2000_2026_v1.json').read_text())['vehicles']
    rows=workbook_rows(args.workbook)
    indices=[]
    with concurrent.futures.ThreadPoolExecutor(max_workers=3) as ex:
        for batch in ex.map(lambda kv:index(*kv),MAKERS.items()): indices.extend(batch)
    print('Weds generation pages:',len(indices),flush=True)
    details=[]
    with concurrent.futures.ThreadPoolExecutor(max_workers=3) as ex:
        for i,result in enumerate(ex.map(detail,indices),1):
            details.append(result)
            if i%100==0: print('Read',i,'/',len(indices),flush=True)
    audit=[]
    for model in master:
        names={norm(model['model']),*(norm(a) for a in model.get('aliases',[]))}
        tr=[r for r in rows if r['maker']==model['maker'] and norm(r['model']) in names]
        wd=[r for r in details if r['maker']==model['maker'] and norm(r['model']) in names]
        audit.append(dict(search_id=model['search_id'],maker=model['maker'],model=model['model'],
                          topy_rows=[r['row'] for r in tr],weds=wd))
    result=dict(workbook=Path(args.workbook).name,workbook_sha256=hashlib.sha256(Path(args.workbook).read_bytes()).hexdigest(),
                searched_models=len(audit),topy=rows,weds=details,audit=audit)
    Path(args.out).write_text(json.dumps(result,ensure_ascii=False,indent=2)+'\n')
    print(json.dumps({'searched':len(audit),'topy_matches':sum(bool(r['topy_rows']) for r in audit),
        'weds_matches':sum(bool(r['weds']) for r in audit),'errors':sum('error' in r for r in details)},ensure_ascii=False))

if __name__=='__main__': main()
